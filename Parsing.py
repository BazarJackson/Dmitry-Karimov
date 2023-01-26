from selenium import webdriver
from selenium.webdriver.firefox.firefox_profile import FirefoxProfile
import time
import csv
from datetime import date
from openpyxl import load_workbook
import pandas as pd
import openpyxl
from openpyxl.writer.excel import save_workbook
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
caps2 = DesiredCapabilities().FIREFOX
caps2["pageLoadStrategy"] = "normal"


def disableImages(self):
    ## get the Firefox profile object
    firefoxProfile = FirefoxProfile()
    ## Disable CSS
    firefoxProfile.set_preference('permissions.default.stylesheet', 2)
    ## Disable images
    firefoxProfile.set_preference('permissions.default.image', 2)
    ## Disable Flash
    firefoxProfile.set_preference('dom.ipc.plugins.enabled.libflashplayer.so', 'false')
    ## Set the modified profile while creating the browser object 
    self.browserHandle = webdriver.Firefox(firefoxProfile)
driver = webdriver.Firefox(desired_capabilities=caps2)
driver.set_window_position(1100, -2)

#КРИСТАЛЬНАЯ
driver.get('https://voda174.ru/')
price = driver.find_elements_by_xpath("/html/body/div[1]/div[2]/div[4]/div[3]/div/div[6]/div")
for item3 in price:
    a = item3.text
    b = ' '.join(a.split(" ")[3:-1])
    price3 = int(b)
    print("Кристальная, цена за 1 шт.", price3)
time.sleep(2)
price = driver.find_elements_by_xpath("/html/body/div[1]/div[2]/div[4]/div[3]/div/div[10]/div/b/span")
for item4 in price:
    a = item4.text
    b = ' '.join(a.split(" ")[:-1])
    price4 = int(b)
    print("Кристальная, цена от 2 шт.", price4)

#ЛЮКС ВОДА
driver.get('https://l-w.ru/catalog/voda/voda_19l/')
but = driver.find_elements_by_class_name('fancybox-close-small')
for i in but:
    i.click()
price = driver.find_elements_by_xpath('/html/body/div[1]/main/div/div/div/div[2]/div[1]/div[1]/div[2]')
for item14 in price:
    a = item14.text
    b = a[:3]
    price14 = int(b)
    print("Люкс вода, цена от 1 шт.", price14)
time.sleep(1)
but = driver.find_elements_by_xpath('//*[@id="content"]/div/div/div/div[2]/div[1]/div[2]/button[2]')
for i in but:
    i.click()
price = driver.find_elements_by_xpath('/html/body/div[1]/main/div/div/div/div[2]/div[1]/div[1]/div[2]')
for item15 in price:
    a = item15.text
    b = a[:3]
    price15 = int(b)
    print("Люкс вода, цена от 2 шт.", price15)


caps = DesiredCapabilities().FIREFOX
caps["pageLoadStrategy"] = "eager"
browser = webdriver.Firefox(desired_capabilities=caps)
browser.set_window_position(-100, -2)

#НИАГАРА
browser.get('https://niagara74.ru/catalog/pitevaya_voda/19_l_niagara_premium_pit_voda/')
price = browser.find_elements_by_xpath('/html/body/div[2]/header/div[1]/div[1]/div[1]/div[2]/a/svg')
for i in price:
    i.click()
price = browser.find_elements_by_xpath('//*[@id="bx_117848907_385_price"]')
for item12 in price:
    a = item12.text
    b = ' '.join(a.split(" ")[:-1])
    price12 = int(b)
    print("Ниагара, цена от 2 шт.", price12)
price = browser.find_elements_by_xpath('//*[@id="bx_117848907_385_old_price"]')
for item13 in price:
    a = item13.text
    b = ' '.join(a.split(" ")[:-1])
    price13 = int(b)
    print("Ниагара, цена за 1 шт.", price13)


#ЖИВАЯ КАПЛЯ
browser.get('https://живаякапля.рф/catalog/water/water_drinking_zhivaya_kaplya_19l/')
time.sleep(2)
but = browser.find_elements_by_xpath('//*[@id="bx_117848907_1012_prop_842_list"]/li[2]/span/span')
for i in but:
    i.click()
time.sleep(2)
price = browser.find_elements_by_xpath("/html/body/div[6]/div[7]/div[2]/div/div/div/div/div/div/div/div[2]/div/div[2]/div[1]/div[1]/div[1]/span[1]")
for item1 in price:
    b = item1.text
    price1 = int(b)
    print("Живая капля, цена за 1 шт.", price1)

time.sleep(2)
but = browser.find_elements_by_xpath('//*[@id="bx_117848907_1012_prop_842_list"]/li[6]/span/span')
for i in but:
    i.click()
but = browser.find_elements_by_xpath('//*[@id="content"]/div[2]/div/div/div/div/div/div/div/div[2]/div/div[5]/div/div[2]/div[1]/div[1]/div/span[2]')
for i in but:
    i.click()
time.sleep(2)
price = browser.find_elements_by_xpath("/html/body/div[6]/div[7]/div[2]/div/div/div/div/div/div/div/div[2]/div/div[2]/div[1]/div[1]/div[1]/span[1]")
for item2 in price:
    b = item2.text
    price2 = int(b)
    print("Живая капля, цена от 2 шт.", price2)


#ГОРНЫЙ ОАЗИС
browser.get('https://www.74mv.ru/rezultat-filtratsii/горный-оазис/?custom_f_3[0]=3139d0bb')

price = browser.find_elements_by_xpath('/html/body/div[1]/main/div[1]/div[2]/div[2]/div/div[2]/div/div[3]/div[1]/div/div[3]/div[1]/div/div[1]/div[2]/span[2]')
for item5 in price:
    b = item5.text
    br = b[:3]
    price5 = int(br)
    print("Горный Оазис (заказ только от 2 шт.), цена от 2 шт.", price5)



#ВЛАСОВ КЛЮЧ
browser.get('http://vlasovkluch.ru/cat/product/item_6.html')
time.sleep(3)
browser.refresh()
price = browser.find_elements_by_xpath('/html/body/div/div[1]/main/div/div/div[1]/div/div[2]/div/div/div[1]/div/div[2]/div[1]/div[3]/div/div[1]')
for item6 in price:
    a = item6.text
    b = ' '.join(a.split(" ")[:-1])
    price6 = int(b)
    print("Власов ключ (нет акции за 2 шт.), цена за 1 шт.", price6)


#ЧЕБАРКУЛЬСКИЙ ИСТОК
browser.get('https://chebistok.ru/')
price = browser.find_elements_by_xpath('/html/body/div[3]/section[6]/div/div[1]/div[2]/div[5]/p[3]')
for item7 in price:
    a = item7.text
    b = ' '.join(a.split(" ")[-2:])
    c = b[:3]
    price7 = int(c)
    print("Чебаркуль, цена за 1 шт.", price7)
price = browser.find_elements_by_xpath('/html/body/div[3]/section[6]/div/div[1]/div[2]/div[5]/p[2]')
for item8 in price:
    a = item8.text
    b = ' '.join(a.split(" ")[:-1])
    price8 = int(b)
    print("Чебаркуль, цена от 2 шт.", price8)


#ЛЮБИМАЯ+
browser.get('https://vodalubima.ru/')
time.sleep(2)
browser.execute_script("window.scrollTo(0, 1500)")
time.sleep(2)
price = browser.find_elements_by_xpath('/html/body/div[1]/div[7]/div/div/div[1]/div/div[2]/div[3]/div/div[1]')
for item9 in price:
    a = item9.text
    price9 = int(a)
    print("Любимая, цена от 2 шт.", price9)


#АРТЕНЗА
browser.get('https://aqua-mobil.ru/voda-9-19-litrov/voda-artezianskaya-artenza-19-l-oborotnaya-tara.html?ysclid=l8y63jyh82317809998')
price = browser.find_elements_by_xpath('/html/body/div/div[1]/div[3]/div/div[2]/div[5]/div[1]/div/form/div[1]/div[1]/div[2]/span')
for item10 in price:
    a = item10.text
    price10 = int(a)
    print("Артенза, цена от 2 шт.", price10)
time.sleep(1)
but = browser.find_elements_by_xpath("/html/body/div/div[1]/div[3]/div/div[2]/div[5]/div[1]/div/form/div[1]/div[2]/div[2]/button[1]")
for i in but:
    i.click()

price = browser.find_elements_by_xpath('/html/body/div/div[1]/div[3]/div/div[2]/div[5]/div[1]/div/form/div[1]/div[1]/div[2]/span')
for item11 in price:
    a = item11.text
    price11 = int(a)
    print("Артенза, цена за 1 шт.", price11)



browser.quit()
driver.quit()

today = date.today()
fn = 'parsing.xlsx'
wb = load_workbook(fn)
ws = wb.active 
b2 = ws['B2']
b3 = ws['B3']
c2 = ws['C2']
c3 = ws['C3']
d3 = ws['D3']
e2 = ws['E2']
e3 = ws['E2']
f2 = ws['F2']
f3 = ws['F3']
g3 = ws['G3']
h2 = ws['H2']
h3 = ws['H3']
i2 = ws['I2']
i3 = ws['I3']
j2 = ws['J2']
j3 = ws['J3']
ws = wb.create_sheet(str(today))
ws.append([' ','ЖИВАЯ КАПЛЯ', 'КРИСТАЛЬНАЯ', 'ГОРНЫЙ ОАЗИС', 'ВЛАСОВ КЛЮЧ', 'ЧЕБАРКУЛЬСКИЙ ИСТОК', 'ЛЮБИМАЯ+', 'АРТЕНЗА', 'НИАГАРА', 'ЛЮКС ВОДА'])
ws.append(['Цена 1шт',price1, price3, " - ", price6, price7, " - ", price11, price13, price14])
ws.append(['Цена от 2шт',price2, price4, price5, price6, price8, price9, price10, price12, price15])
ws.append(['Изменение цены с прошлого проверки', str(round(((price1 - b2.value)/price1) * 100, 2))+"%", str(round(((price3 - c2.value)/price3)*100, 2))+"%", ' ', str(round(((price6 - e2.value)/price6)*100, 2))+"%", str(round(((price7 - f2.value)/price7)*100, 2))+"%", ' ', str(round(((price11 - h2.value)/price11)*100, 2))+"%", str(round(((price13 - i2.value)/price13)*100, 2))+"%", str(round(((price14 - j2.value)/price14)*100, 2))+"%"])
ws.append(['Изменение цены с прошлой проверки', str(round(((price2 - b3.value)/price2) * 100, 2))+"%", str(round(((price4 - c3.value)/price4)*100, 2))+"%", str(round(((price5 - d3.value)/price5)*100, 2))+"%", str(round(((price6 - e3.value)/price6)*100, 2))+"%", str(round(((price8 - f3.value)/price8)*100, 2))+"%", str(round(((price9 - g3.value)/price9)*100, 2))+"%", str(round(((price10 - h3.value)/price10)*100, 2))+"%", str(round(((price12 - i3.value)/price12)*100, 2))+"%", str(round(((price15 - j3.value)/price15)*100, 2))+"%"])
wb.save(fn)
wb.close()
print("Excel фаил загружен")




