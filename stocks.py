from datetime import date
from openpyxl import load_workbook
import pandas as pd
import openpyxl
from openpyxl.writer.excel import save_workbook
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium import webdriver

caps2 = DesiredCapabilities().FIREFOX
caps2["pageLoadStrategy"] = "normal"
driver = webdriver.Firefox(desired_capabilities=caps2)
driver.set_window_position(800, -2)
#Люкс Вода "50% скидка" (1)
driver.get('https://l-w.ru/promo/')
but = driver.find_elements_by_class_name('fancybox-close-small')
for i in but:
    i.click()
sale = driver.find_elements_by_xpath('//*[@id="bx_3218110189_44293"]/div[2]')
for i in sale:
    sale11 = i.text
    print(sale11)
print("--------------------------------------------------------------------------------------------------------------------------------")
#Люкс вода "спасибо мама" (2)

sale = driver.find_elements_by_xpath('//*[@id="bx_3218110189_1607"]/div[2]')
for i in sale:
    sale12 = i.text
    print(sale12)
print("--------------------------------------------------------------------------------------------------------------------------------")
#Люкс вода "бонусы по картам" (3)

sale = driver.find_elements_by_xpath('//*[@id="bx_3218110189_1671"]/div[2]')
for i in sale:
    sale13 = i.text
    print(sale13)
print("--------------------------------------------------------------------------------------------------------------------------------")
#Люкс вода "кулер в подарок" (4)

sale = driver.find_elements_by_xpath('//*[@id="bx_3218110189_1609"]/div[2]')
for i in sale:
    sale14 = i.text
    print(sale14)
print("--------------------------------------------------------------------------------------------------------------------------------")
#Люкс вода "15% в приложении" (5)

sale = driver.find_elements_by_xpath('//*[@id="bx_3218110189_50318"]/div[2]')
for i in sale:
    sale15 = i.text
    print(sale15)
print("--------------------------------------------------------------------------------------------------------------------------------")
#Люкс вода "пустые на полные" (6)
driver.execute_script("window.scrollTo(0, 1050)") 
sale = driver.find_elements_by_xpath('//*[@id="bx_3218110189_50182"]/div[2]')
for i in sale:
    sale16 = i.text
    print(sale16)

print("--------------------------------------------------------------------------------------------------------------------------------")
#Люкс вода "ПЭТ не мусор" (7)

sale = driver.find_elements_by_xpath('//*[@id="bx_3218110189_43025"]/div[2]')
for i in sale:
    sale17 = i.text
    print(sale17)
print("--------------------------------------------------------------------------------------------------------------------------------")
#Люкс вода "10% на сайте" (8)

sale = driver.find_elements_by_xpath('//*[@id="bx_3218110189_51713"]/div[2]')
for i in sale:
    sale18 = i.text
    print(sale18)
print("--------------------------------------------------------------------------------------------------------------------------------")
#Люкс вода "почетный возраст" (9)

sale = driver.find_elements_by_xpath('//*[@id="bx_3218110189_33156"]/div[2]')
for i in sale:
    sale19 = i.text
    print(sale19)
print("--------------------------------------------------------------------------------------------------------------------------------")
#Люкс вода "предложение СТАРТ" (10)

sale = driver.find_elements_by_xpath('//*[@id="bx_3218110189_25733"]/div[2]')
for i in sale:
    sale110 = i.text
    print(sale110)
print("--------------------------------------------------------------------------------------------------------------------------------")


driver.get('https://l-w.ru/promo/?PAGEN_1=2')
#Люкс вода "кулер в рассрочку" (11)

sale = driver.find_elements_by_xpath('//*[@id="bx_3218110189_22061"]/div[2]')
for i in sale:
    sale111 = i.text
    print(sale111)
print("--------------------------------------------------------------------------------------------------------------------------------")
#Люкс вода "доставка в школы" (12)
sale = driver.find_elements_by_xpath('//*[@id="bx_3218110189_15973"]/div[2]')
for i in sale:
    sale112 = i.text
    print(sale112)
print("--------------------------------------------------------------------------------------------------------------------------------")
#Люкс вода "вода в подарок" (13)

sale = driver.find_elements_by_xpath('//*[@id="bx_3218110189_1610"]/div[2]')
for i in sale:
    sale113 = i.text
    print(sale113)
print("--------------------------------------------------------------------------------------------------------------------------------")

driver.get('https://chebistok.ru/#sale')
with open('Чебаркуль.png', 'wb') as file:
   l = driver.find_element_by_xpath('/html/body/div[3]/section[2]/div/div/div[2]/div[3]/div[1]/img[1]')
   file.write(l.screenshot_as_png)
print('картинка 1 загружена')

element = driver.find_element_by_xpath('/html/body/div[3]/section[2]/div/div/div[2]/div[2]')
element.click()

with open('Чебаркуль2.png', 'wb') as file:
   l = driver.find_element_by_xpath('/html/body/div[3]/section[2]/div/div/div[2]/div[3]/div[2]/img[1]')
   file.write(l.screenshot_as_png)
print('картинка 2 загружена')
driver.get('https://aqua-mobil.ru/aktsii/aktsii_16.html')
sale = driver.find_elements_by_xpath('/html/body/div/div[1]/div[3]/div/div[2]/div/div[3]/strong')
for i in sale:
    sale55 = i.text
    print(i.text)
with open('Аквамобиль.png', 'wb') as file:
   l = driver.find_element_by_xpath('/html/body/div/div[1]/div[3]/div/div[2]/div/div[5]/img')
   file.write(l.screenshot_as_png)
print('картинка загружена')
#Niagara---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

caps = DesiredCapabilities().FIREFOX
caps["pageLoadStrategy"] = "eager"
browser = webdriver.Firefox(desired_capabilities=caps)
browser.set_window_position(-100, -2)
#Niagara "старт набор" (1)
browser.get('https://niagara74.ru/stock/aktsiya-dlya-novykh-klientov-2-vody-niagara-19l-pompa-za-399-rubley/')
sale = browser.find_elements_by_xpath('//*[@id="main"]/div/div/div/div/div[2]/div/div[2]')
for i in sale:
    sale21 = i.text
    print(sale21)
print("--------------------------------------------------------------------------------------------------------------------------------")
#Niagara "в школы и детсады" (1)
browser.get('https://niagara74.ru/stock/skidki-do-40-dlya-detskikh-sadov-i-shkol/')
sale = browser.find_elements_by_xpath('//*[@id="main"]/div/div/div/div/div[2]/div/div[2]')
for i in sale:
    sale22 = i.text
    print(sale22)
print("--------------------------------------------------------------------------------------------------------------------------------")
#Niagara "для пенсионеров" (1)
browser.get('https://niagara74.ru/stock/lgotnye-tseny-dlya-pensionerov/')
sale = browser.find_elements_by_xpath('//*[@id="main"]/div/div/div/div/div[2]/div/div[2]')
for i in sale:
    sale23 = i.text
    print(sale23)
print("--------------------------------------------------------------------------------------------------------------------------------")
#Niagara "тара в залог" (1)
browser.get('https://niagara74.ru/stock/polikarbonatnaya-tara-v-zalog/')
sale = browser.find_elements_by_xpath('//*[@id="main"]/div/div/div/div/div[2]/div/div[2]')
for i in sale:
    sale24 = i.text
    print(sale24)
print("--------------------------------------------------------------------------------------------------------------------------------")
#Niagara "бонуска" (1)
browser.get('https://niagara74.ru/stock/bonusnaya-programma-/')
sale = browser.find_elements_by_xpath('//*[@id="main"]/div/div/div/div/div[2]/div/div[2]')
for i in sale:
    sale25 = i.text
    print(sale25)
print("--------------------------------------------------------------------------------------------------------------------------------")
#Niagara "кулер в подарок" (1)
browser.get('https://niagara74.ru/stock/kuler-v-podarok/')
sale = browser.find_elements_by_xpath('//*[@id="main"]/div/div/div/div/div[2]/div/div[2]')
for i in sale:
    sale26 = i.text
    print(sale26)
print("--------------------------------------------------------------------------------------------------------------------------------")
#Niagara "получи пиццу" (1)
browser.get('https://niagara74.ru/stock/nachni-pit-niagara-i-poluchi-srednyuyu-dodo-pitstsu-v-podarok/')
sale = browser.find_elements_by_xpath('//*[@id="main"]/div/div/div/div/div[2]/div/div[2]')
for i in sale:
    sale27 = i.text
    print(sale27)
print("--------------------------------------------------------------------------------------------------------------------------------")

#Niagara "20 шт 840 руб" (1)
browser.get('https://niagara74.ru/stock/20-sht-n-5l-778-rub-i-20-np-5-lit-838-rub-samovyvoz/')
sale = browser.find_elements_by_xpath('//*[@id="main"]/div/div/div/div/div[2]/div/div[2]')
for i in sale:
    sale29 = i.text
    print(sale29)
print("--------------------------------------------------------------------------------------------------------------------------------")
#Niagara "9+1" (1)
browser.get('https://niagara74.ru/stock/9-1-dlya-vsekh-pokupateley-samovyvoz/')
sale = browser.find_elements_by_xpath('//*[@id="main"]/div/div/div/div/div[2]/div/div[2]')
for i in sale:
    sale210 = i.text
    print(sale210)
print("--------------------------------------------------------------------------------------------------------------------------------")

#Живая капля "5% от стоимости" (31) -----------------------------------------------------------------------------------------------------------------------------------------------------------------------
browser.get('https://живаякапля.рф/sale/vernem_5_ot_stoimosti_pokupki/')
sale = browser.find_elements_by_xpath('//*[@id="content"]/div[2]/div/div/div/div/div[1]/div/div[3]/div/div')
for i in sale:
    sale31 = i.text
    print(sale31)
print("--------------------------------------------------------------------------------------------------------------------------------")

#Живая капля "сравни качество" (32)
browser.get('https://живаякапля.рф/sale/sravni_kachestvo/')
sale = browser.find_elements_by_xpath('//*[@id="content"]/div[2]/div/div/div/div/div[1]/div/div[3]/div/div')
for i in sale:
    sale32 = i.text
    print(sale32)

print("--------------------------------------------------------------------------------------------------------------------------------")

#Живая капля "новогодняя" (33)
browser.get('https://живаякапля.рф/sale/novogodniy_perepolokh/')
sale = browser.find_elements_by_xpath('//*[@id="content"]/div[2]/div/div/div/div/div[1]/div/div[3]/div[2]/div')
for i in sale:
    sale33 = i.text
    print(sale33)

print("--------------------------------------------------------------------------------------------------------------------------------")

#Живая капля "выгодный старт" (34)
browser.get('https://живаякапля.рф/sale/vygodnyy_start_ot_777/')
sale = browser.find_elements_by_xpath('//*[@id="content"]/div[2]/div/div/div/div/div[1]/div/div[3]/div/div')
for i in sale:
    sale34 = i.text
    print(sale34)

print("--------------------------------------------------------------------------------------------------------------------------------")

#Горный оазис -----------------------------------------------------------------------------------------------------------------------------------------------------------------------

browser.get('https://www.74mv.ru/aktsii')
sale = browser.find_elements_by_xpath('//*[@id="bd_results"]/div[2]/div/div[1]/p[1]')
for i in sale:
    sale41 = i.text
    print(i.text)
sale = browser.find_elements_by_xpath('//*[@id="bd_results"]/div[2]/div/div[2]/p[1]')
for i in sale:
    sale42 = i.text
    print(i.text)
sale = browser.find_elements_by_xpath('//*[@id="bd_results"]/div[2]/div/div[3]/p[1]')
for i in sale:
    sale43 = i.text
    print(i.text)
sale = browser.find_elements_by_xpath('//*[@id="bd_results"]/div[2]/div/div[4]/p[1]')
for i in sale:
    sale44 = i.text
    print(i.text)

#Аквамобиль -----------------------------------------------------------------------------------------------------------------------------------------------------------------------
print('АКВА МОБИЛЬ')
browser.get('https://aqua-mobil.ru/aktsii/aktsii_14.html')
sale = browser.find_elements_by_xpath('/html/body/div/div[1]/div[3]/div/div[2]/div/p')
for i in sale:
    sale51 = i.text
    print(i.text)
sale = browser.find_elements_by_xpath('/html/body/div/div[1]/div[3]/div/div[2]/div/h3')
for i in sale:
    sale52 = i.text
    print(i.text)
 
browser.get('https://aqua-mobil.ru/aktsii/aktsii_10.html')
sale = browser.find_elements_by_xpath('/html/body/div/div[1]/div[3]/div/div[2]/div')
for i in sale:
    sale53 = i.text
    print(i.text)
 
browser.get('https://aqua-mobil.ru/aktsii/aktsii_11.html')
sale = browser.find_elements_by_xpath('/html/body/div/div[1]/div[3]/div/div[2]/div')
for i in sale:
    sale54 = i.text
    print(i.text)
 



#-------------------------------------------------- КАРТИНКИ

browser.close()
driver.close()
fn = 'proj.xlsx'
wb = load_workbook(fn)
ws = wb.active 
ws['A1'] = 'Люкс вода'
ws['A2'] = sale11
ws['A3'] = sale12
ws['A4'] = sale13
ws['A5'] = sale14
ws['A6'] = sale15
ws['A7'] = sale16
ws['A8'] = sale17
ws['A9'] = sale18
ws['A10'] = sale19
ws['A11'] = sale110
ws['A12'] = sale111
ws['A13'] = sale112
ws['A14'] = sale113


ws['B1'] = 'Niagara'
ws['B2'] = sale21
ws['B3'] = sale22
ws['B4'] = sale23
ws['B5'] = sale24
ws['B6'] = sale25
ws['B7'] = sale26
ws['B8'] = sale27
ws['B9'] = sale29
ws['B10'] = sale210

ws['C1'] = 'Живая капля'
ws['C2'] = sale31
ws['C3'] = sale32
ws['C4'] = sale33
ws['C5'] = sale34

ws['D1'] = 'Горный оазис'
ws['D2'] = sale41
ws['D3'] = sale42
ws['D4'] = sale43
ws['D5'] = sale44

ws['E1'] = 'Аквамобиль'
ws['E2'] = sale51
ws['E3'] = sale52
ws['E4'] = sale53
ws['E5'] = sale54
ws['E6'] = sale55


wb.save(fn)
wb.close()